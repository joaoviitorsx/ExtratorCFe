import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.models.cfeModel import CFeModel

class ExportService:
    def __init__(self):
        self.header_style = self.estiloCabecalho()

    def gerarPlanilha(self, cfes_dict: dict, caminho_saida: str) -> str:
        wb = Workbook()
        wb.remove(wb.active)

        self.abaVendas(wb, cfes_dict)
        self.abaCancelados(wb, cfes_dict)
        self.abaForaPadrao(wb, cfes_dict)

        if not caminho_saida.lower().endswith(".xlsx"):
            caminho_saida += ".xlsx"

        os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
        wb.save(caminho_saida)
        return caminho_saida

    def abaVendas(self, wb, cfes_dict):
        vendas = cfes_dict.get("venda_validada", []) + cfes_dict.get("venda_presat", [])
        if vendas:
            self.abaCompleta(wb, "Vendas", vendas)

    def abaCancelados(self, wb, cfes_dict):
        cancelados = cfes_dict.get("cancelamento_validado", []) + cfes_dict.get("cancelamento_presat", [])
        if cancelados:
            self.abaCompleta(wb, "Cancelados", cancelados)

    def abaForaPadrao(self, wb, cfes_dict):
        fora_padrao = cfes_dict.get("fora_padrao", [])
        if fora_padrao:
            ws = wb.create_sheet("Fora do Padrão")
            headers = ["Arquivo", "Motivo"]
            ws.append(headers)
            self.formatarCabecalho(ws, headers)
            for arquivo in fora_padrao:
                ws.append([arquivo, "Arquivo inválido ou não é um XML de CF-e"])


    def abaCompleta(self, wb: Workbook, nome_aba: str, lista_cfe: list):
        ws = wb.create_sheet(nome_aba)
        headers = self.montarHeaders()
        ws.append(headers)
        self.formatarCabecalho(ws, headers)

        linhas = []
        for cfe in lista_cfe:
            linhas.extend(self.linhasCfe(cfe))
        for linha in linhas:
            ws.append(linha)

    def montarHeaders(self):
        return [
            "Chave", "Versão CF-e", "Versão Dados Ent", "Versão SAT",
            "Código UF", "Código Numérico", "Modelo", "Nº Série SAT", "Nº CF-e",
            "Data Emissão", "Hora Emissão", "Dígito Verificador", "Tipo Ambiente",
            "CNPJ Software House", "Assinatura AC", "Assinatura QR Code", "Número do Caixa",
            "CNPJ Emitente", "Nome Social", "Nome Fantasia",
            "Logradouro Emitente", "Número Emitente", "Complemento Emitente",
            "Bairro Emitente", "Município Emitente", "CEP Emitente",
            "IE Emitente", "Regime Tributário", "Indicador ISSQN",
            "CPF/CNPJ Destinatário", "Nome Destinatário",
            "Código Produto", "Descrição Produto", "EAN", "NCM", "CEST", "CFOP",
            "Unidade", "Quantidade", "V. Unitário", "V. Total", "Desconto", "Outros", "Regra",
            "Valor Tributos Lei 12741", "ICMS Origem", "ICMS CST", "ICMS Alíquota", "ICMS Valor",
            "PIS CST", "PIS Base", "PIS Alíquota", "PIS Valor",
            "COFINS CST", "COFINS Base", "COFINS Alíquota", "COFINS Valor",
            "Total ICMS", "Total Produtos", "Total Desconto", "Total PIS", "Total COFINS",
            "Total CF-e", "Total Tributos Lei 12741",
            "Forma Pagamento", "Valor Pago", "Troco",
            "Observações do Fisco", "Informações Complementares"
        ]

    def linhasCfe(self, cfe: CFeModel):
        ide, emit, dest = cfe.ide, cfe.emitente, cfe.destinatario
        totais, pagamentos = cfe.totais, cfe.pagamentos
        infAdic = cfe.infAdic
        obsFisco = " | ".join([f"{o.get('xCampo', '')}: {o.get('xTexto', '')}" for o in cfe.obsFisco]) if cfe.obsFisco else "-"

        forma_pagamento = " | ".join([p.get('cMP', '-') for p in pagamentos]) if pagamentos else "-"
        valor_pago = " | ".join([p.get('vMP', '-') for p in pagamentos]) if pagamentos else "-"
        troco = totais.get("vTroco", "-") if "vTroco" in totais else "-"

        linhas = []
        for item in cfe.itens:
            icms = self.extrairImposto(item.impostos, "ICMS")
            pis = self.extrairImposto(item.impostos, "PIS")
            cofins = self.extrairImposto(item.impostos, "COFINS")

            linha = [
                cfe.chave, cfe.versao, cfe.versaoDadosEnt, cfe.versaoSB,
                ide.get("cUF", "-"), ide.get("cNF", "-"), ide.get("mod", "-"),
                ide.get("nserieSAT", "-"), ide.get("nCFe", "-"),
                ide.get("dEmi", "-"), ide.get("hEmi", "-"), ide.get("cDV", "-"),
                ide.get("tpAmb", "-"), ide.get("CNPJ", "-"),
                ide.get("signAC", "-"), cfe.assinaturaQRCODE, ide.get("numeroCaixa", "-"),

                emit.get("CNPJ", "-"), emit.get("xNome", "-"), emit.get("xFant", "-"),
                emit.get("enderEmit", {}).get("xLgr", "-"), emit.get("enderEmit", {}).get("nro", "-"),
                emit.get("enderEmit", {}).get("xCpl", "-"), emit.get("enderEmit", {}).get("xBairro", "-"),
                emit.get("enderEmit", {}).get("xMun", "-"), emit.get("enderEmit", {}).get("CEP", "-"),
                emit.get("IE", "-"), emit.get("cRegTrib", "-"), emit.get("indRatISSQN", "-"),

                dest.get("CPF", dest.get("CNPJ", "-")), dest.get("xNome", "-"),

                item.cProd, item.xProd, item.cEAN, item.NCM, item.CEST,
                item.CFOP, item.uCom, item.qCom, item.vUnCom, item.vProd,
                item.vDesc, item.vOutro, item.indRegra,

                item.vItem12741,
                icms.get("Orig", "-"), icms.get("CST", "-"), icms.get("pICMS", "-"), icms.get("vICMS", "-"),
                pis.get("CST", "-"), pis.get("vBC", "-"), pis.get("pPIS", "-"), pis.get("vPIS", "-"),
                cofins.get("CST", "-"), cofins.get("vBC", "-"), cofins.get("pCOFINS", "-"), cofins.get("vCOFINS", "-"),

                totais.get("ICMSTot", {}).get("vICMS", "-"),
                totais.get("ICMSTot", {}).get("vProd", "-"),
                totais.get("ICMSTot", {}).get("vDesc", "-"),
                totais.get("ICMSTot", {}).get("vPIS", "-"),
                totais.get("ICMSTot", {}).get("vCOFINS", "-"),
                totais.get("vCFe", "-"), totais.get("vCFeLei12741", "-"),

                forma_pagamento, valor_pago, troco,
                obsFisco, infAdic.get("infCpl", "-")
            ]
            linhas.append(linha)
        return linhas

    def estiloCabecalho(self):
        return {"font": Font(bold=True, color="FFFFFF"),
                "fill": PatternFill("solid", fgColor="4F81BD"),
                "align": Alignment(horizontal="center")}

    def formatarCabecalho(self, ws, headers):
        for cell in ws[1]:
            cell.font = self.header_style["font"]
            cell.fill = self.header_style["fill"]
            cell.alignment = self.header_style["align"]
        for i, header in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = max(len(header) + 2, 12)

    def extrairImposto(self, impostos: dict, tipo: str) -> dict:
        if tipo not in impostos:
            return {}
        bloco = impostos[tipo]
        if isinstance(bloco, dict):
            for chave, dados in bloco.items():
                if isinstance(dados, dict):
                    return dados
        return {}
